import streamlit as st
st.set_page_config(page_title="Vyhodnocení laboratorního deníku")
st.write("Streamlit import OK")
import pandas as pd
st.write("Pandas import OK")
import io
st.write("io import OK")
from openpyxl import load_workbook
st.write("openpyxl import OK")

from difflib import SequenceMatcher

def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

def contains_similar(text, keyword, threshold=0.4):
    text = text.lower()
    keyword = keyword.lower()
    if keyword in text:
        return True
    return similar(text, keyword) >= threshold
from difflib import SequenceMatcher

st.title("Vyhodnocení laboratorního deníku")

lab_file = st.file_uploader("Nahraj laboratorní deník (Evidence zkoušek zhotovitele)", type="xlsx")
xlsx_file = st.file_uploader("Nahraj soubor Klíč.xlsx", type="xlsx")

def count_matches_advanced(df, konstrukce, zkouska_raw, stanice_raw):
    druhy_zk = [z.strip().lower() for z in str(zkouska_raw).split(",") if z.strip()]
    staniceni = [s.strip().lower() for s in str(stanice_raw).split(",") if s.strip()]
    match_count = 0

    for index, row in df.iterrows():
        st.markdown(f"📄 **Řádek {index + 2}**: " + " | ".join(str(v) for v in row.values if pd.notna(v)))
        text_row = " ".join(str(v).lower() for v in row.values if pd.notna(v))
        konstrukce_ok = contains_similar(text_row, konstrukce)
        stanice_ok = any(s in text_row for s in staniceni)
        zkouska_ok = any(z in text_row for z in druhy_zk)

    return match_count


def process_op_sheet(key_df, target_df):
    if "D" not in target_df.columns:
        target_df["D"] = 0
    if "E" not in target_df.columns:
        target_df["E"] = ""
    for i in range(1, len(target_df)):
        row = target_df.iloc[i]
        zasyp = str(row.iloc[0])
        if pd.isna(zasyp):
            continue
        matches = key_df[key_df.iloc[:, 0] == zasyp]
        total_count = 0
        for _, mrow in matches.iterrows():
            konstrukce = mrow.get("konstrukční prvek", "")
            zkouska = mrow.get("druh zkoušky", "")
            stanice = mrow.get("staničení", "")
            if konstrukce and zkouska and stanice:
                total_count += count_matches_advanced(lab_df, konstrukce, zkouska, stanice)
        target_df.at[i, "D"] = total_count
        pozadovano = row.get("C")
        if pd.notna(pozadovano):
            target_df.at[i, "E"] = "Vyhovující" if total_count >= pozadovano else f"Chybí {abs(int(pozadovano - total_count))} zk."
    return target_df


def process_cely_objekt_sheet(key_df, target_df):
    for i, row in target_df.iterrows():
        material = row.get("materiál")
        zkouska = row.get("druh zkoušky")
        if pd.isna(material) or pd.isna(zkouska):
            continue
        count = count_matches_advanced(lab_df, material, zkouska, "")
        target_df.at[i, "C"] = count
        pozadovano = row.get("B")
        if pd.notna(pozadovano):
            target_df.at[i, "D"] = "Vyhovující" if count >= pozadovano else f"Chybí {abs(int(pozadovano - count))} zk."
    return target_df

if lab_file and xlsx_file:
    lab_bytes = lab_file.read()
    lab_df = pd.read_excel(io.BytesIO(lab_bytes), sheet_name="Evidence zkoušek zhotovitele")

    try:
        xlsx_bytes = xlsx_file.read()
        workbook = load_workbook(io.BytesIO(xlsx_bytes))

        def load_sheet_df(name):
            return pd.read_excel(io.BytesIO(xlsx_bytes), sheet_name=name)

        sheet_names = workbook.sheetnames

        def sheet_exists(name):
            return name in sheet_names

        op1_key = load_sheet_df("seznam zkoušek PM+LM OP1") if sheet_exists("seznam zkoušek PM+LM OP1") else pd.DataFrame()
        op2_key = load_sheet_df("seznam zkoušek PM+LM OP2") if sheet_exists("seznam zkoušek PM+LM OP2") else pd.DataFrame()
        cely_key = load_sheet_df("seznam zkoušek Celý objekt") if sheet_exists("seznam zkoušek Celý objekt") else pd.DataFrame()

        sheet_targets = [
            ("PM - OP1", op1_key),
            ("LM - OP1", op1_key),
            ("PM - OP2", op2_key),
            ("LM - OP2", op2_key),
            ("Celý objekt", cely_key),
        ]

        for sheet_name, key_df in sheet_targets:
            if sheet_exists(sheet_name) and not key_df.empty:
                df = load_sheet_df(sheet_name)
                processed = process_cely_objekt_sheet(key_df, df) if "Celý objekt" in sheet_name else process_op_sheet(key_df, df)
                ws = workbook[sheet_name]
                for i, row in processed.iterrows():
                    if "D" in processed.columns:
                        ws.cell(row=i+2, column=4, value=row.get("D"))
                    if "E" in processed.columns:
                        ws.cell(row=i+2, column=5, value=row.get("E"))

        output = io.BytesIO()
        workbook.save(output)

        st.success("Vyhodnocení dokončeno. Stáhni výsledný soubor níže.")
        st.download_button(
            label="📥 Stáhnout výsledný Excel",
            data=output.getvalue(),
            file_name="vyhodnoceni_vystup.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"Chyba při zpracování souboru: {e}")
